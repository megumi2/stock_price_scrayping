import datetime
import requests as re
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
import time
import sys
import openpyxl
from openpyxl.styles import PatternFill
import os
#BeautifulSoup のバグの修正
sys.path.append("lib.bs4")

#株価の価格や価格推移の情報を格納するためのリスト
price_list = []
price_change_list = []
company_name_list = []

FILEPATH = "output.csv"

INTERVAL_TIME = 10

#リストにスクレイピングしたい企業の証券コードを入力
COMPANY_CODE_LIST = ["1491","1861","1833","1949"]

if __name__ == '__main__':

    #webドライバーの起動
    driver = webdriver.Chrome()

    for company_code in COMPANY_CODE_LIST:   


    
        #URLを作成
        STOCK_URL = 'https://shikiho.toyokeizai.net/stocks/' + company_code
        driver.get(STOCK_URL)
        url = driver.page_source.encode('utf-8')
        #BeautifulSoupで情報を取得
        soup = BeautifulSoup(url, "html.parser")
        company_name = soup.find("div", class_="head__main__left__title").text.strip()
        price = soup.find("div", class_="stock-index__price__current").text.strip()
        price_change = soup.find("div", class_="stock-index__price__change").text.strip()

        #リストに情報を格納
        price_list.append(price)
        price_change_list.append(price_change)
        company_name_list.append(company_name)
        
        #問題なく情報が取得できているかテスト用のコード
        print(company_name +":" + price + ":" + price_change)
        #スクレイピングの時間を空ける
        time.sleep(INTERVAL_TIME)
    #Webドライバーを閉じる
    driver.quit()

    df =pd.DataFrame(data = {'会社名':company_name_list,
                                '株価':price_list,
                                '株価推移':price_change_list
                                })
    df = df.set_index('会社名')
    #ブックを作成
    
    #日程を取得
    date = datetime.datetime.now()
    # ファイルが存在しない場合、新しいExcelファイルを作成
    if not os.path.exists(f'株価リスト_{date.month}.xlsx'):
        with pd.ExcelWriter(f'株価リスト_{date.month}月.xlsx', engine='openpyxl',mode='w') as writer:
            df.to_excel(writer, sheet_name=f'{date.month}月{date.day}日')
    else:
    # ファイルが存在する場合、既存のExcelファイルを読み込んでデータフレームを追加
        with pd.ExcelWriter(f'株価リスト_{date.month}月.xlsx', engine='openpyxl',mode='a') as writer:
            book = openpyxl.load_workbook(f'株価リスト_{date.month}月.xlsx')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            if f'{date.month}月{date.day}日' not in writer.sheets:
                # シートが存在しない場合、新しいシートを作成してデータフレームを追加
                df.to_excel(writer, sheet_name=f'{date.month}月{date.day}日')
            else:
                # シートが既に存在する場合、既存のシートにデータフレームを追加
                writer.sheets[f'{date.month}月{date.day}日'] = writer.book[f'{date.month}月{date.day}日']
                df.to_excel(writer, sheet_name=f'{date.month}月{date.day}日')
            writer.save()
            writer.close()

    #株価リストを編集
    book = openpyxl.load_workbook(f'株価リスト_{date.month}月.xlsx')
    ws = book.active

    red_fill = PatternFill(patternType="solid", fgColor="ff7f50")
    blue_fill = PatternFill(patternType="solid", fgColor="6495ed")
    for rows in ws.iter_rows(min_row=1, min_col=3, max_row=len(company_name_list)+1, max_col=3):
        for cell in rows:
            if "+" in cell.value:
                #セルの背景色を赤くする
                ws[cell.coordinate].fill = red_fill
            elif "-" in cell.value:
                #セルの背景色を青くする
                ws[cell.coordinate].fill = blue_fill
                
    book.save(f'株価リスト_{date.month}月.xlsx')